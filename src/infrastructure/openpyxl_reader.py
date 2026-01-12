from contextlib import contextmanager
from typing import Iterable, Sequence, Any

import openpyxl


class OpenPyxlExcelReader:
    def iter_rows(self, source_path: str) -> Iterable[Sequence[Any]]:
        with self._workbook(source_path) as sheet:
            for row in sheet.iter_rows(values_only=True):
                yield row

    @contextmanager
    def _workbook(self, source_path: str):
        wb = None
        try:
            wb = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
            sheet = wb.active
            if sheet is None:
                raise ValueError("В файле нет активного листа")
            yield sheet
        finally:
            if wb is not None:
                wb.close()
