import os
from datetime import datetime
from typing import Sequence, Any

import openpyxl
from openpyxl.styles import Font, Alignment


class OpenPyxlExcelWriter:
    def write_table(
        self,
        target_path: str,
        headers: Sequence[str],
        rows: Sequence[Sequence[Any]],
        *,
        generated_at_iso: str,
        source_type_label: str = "Excel файл",
        sheet_title: str = "Отфильтрованные данные",
    ) -> None:
        os.makedirs(os.path.dirname(target_path) or ".", exist_ok=True)

        wb = openpyxl.Workbook()
        try:
            ws = wb.active
            ws.title = sheet_title

            ws["A1"] = "Тип файла:"
            ws["A1"].font = Font(bold=True)
            ws["B1"] = source_type_label

            ws["A2"] = "Дата формирования"
            ws["A2"].font = Font(bold=True)
            ws["B2"] = self._format_generated_at(generated_at_iso)

            start_row = 4

            self._write_header(ws, start_row, headers)
            self._write_rows(ws, start_row + 1, rows)
            self._auto_width(ws, headers)

            wb.save(target_path)
        finally:
            wb.close()

    def _format_generated_at(self, generated_at_iso: str) -> str:
        try:
            dt = datetime.fromisoformat(generated_at_iso)
            return dt.strftime("%d.%m.%Y %H:%M")
        except Exception:
            return generated_at_iso

    def _write_header(self, ws, row_idx: int, headers: Sequence[str]) -> None:
        bold = Font(bold=True)
        align = Alignment(horizontal="center", vertical="center")

        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=h)
            cell.font = bold
            cell.alignment = align

    def _write_rows(self, ws, start_row: int, rows: Sequence[Sequence[Any]]) -> None:
        r = start_row
        for row in rows:
            for c, value in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=value)
            r += 1

    def _auto_width(self, ws, headers: Sequence[str]) -> None:
        for col_idx, header in enumerate(headers, start=1):
            letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[letter].width = max(len(str(header)) + 5, 12)
